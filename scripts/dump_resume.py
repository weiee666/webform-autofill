#!/usr/bin/env python3
"""
Dump the maintained personal-info Excel to a structured JSON the skill can read.

Reads:  the resume Excel pointed to by the RESUME_XLSX env var
        (falls back to the maintainer's local default if unset).
Writes: <skill_dir>/cache/resume.json   (next to this script's parent)

Layout of the JSON (one top-level object per sheet, but we only export the two
production sheets):

{
  "zh": {
    "基本信息": {"姓名": "...", "邮箱": "...", ...},
    "教育经历": [{...}, {...}],
    "实习经历": [{...}, {...}],
    "工作经历": [{...}],
    "项目经历": [{...}, ...],
    "校园经历": [...],
    "奖励情况": {...},   # flat key-value if every label only appears once
    "技能栈": {...},
    "链接": {...},
    "_singletons": {label: value, ...}   # everything that's a one-row category
  },
  "en": { ... same shape ... }
}

The script swallows the existing duplicate-N bug by treating each contiguous
block as one record. It also normalises datetime / serial-number dates to
"YYYY.MM" where possible.
"""
import json
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    venv_dir = os.path.join(tempfile.gettempdir(), "claude_openpyxl_venv")
    if not os.path.exists(venv_dir):
        subprocess.check_call([sys.executable, "-m", "venv", venv_dir])
    pip = os.path.join(venv_dir, "bin", "pip")
    subprocess.check_call([pip, "install", "openpyxl", "-q"])
    site_out = subprocess.check_output([
        os.path.join(venv_dir, "bin", "python3"),
        "-c", "import site; print(site.getsitepackages()[0])"
    ]).decode().strip()
    sys.path.insert(0, site_out)
    import openpyxl

# Source Excel path. Override with the RESUME_XLSX env var; otherwise falls back
# to the maintainer's local default. (Keep personal paths out of version control.)
EXCEL_PATH = os.environ.get(
    "RESUME_XLSX",
    os.path.expanduser("~/Desktop/找工作/简历制作/简历信息/网填简历详细版（英文版）_AI维护.xlsx"),
)
if not os.path.exists(EXCEL_PATH):
    raise SystemExit(
        f"找不到 Excel: {EXCEL_PATH}\n"
        "请设置环境变量 RESUME_XLSX 指向你的简历 Excel，例如:\n"
        '  export RESUME_XLSX="/path/to/your_resume.xlsx"'
    )
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE_DIR = os.path.join(SKILL_DIR, "cache")
os.makedirs(CACHE_DIR, exist_ok=True)
OUT_PATH = os.path.join(CACHE_DIR, "resume.json")

SHEET_MAP = {
    "zh": "AI产品经理（中文）",
    "en": "AI产品经理（英文）",
}

# Categories that follow the "类目N" repeated-block pattern.
REPEATED_PATTERN = re.compile(r"^(.+?)\s*([0-9０-９]+)\s*$")
REPEATED_CATEGORIES = {
    "教育经历", "实习经历", "工作经历", "项目经历",
    "校园经历", "亲属关系",
}


def fullwidth_to_int(s: str) -> int:
    out = []
    for c in s:
        if "０" <= c <= "９":
            out.append(str(ord(c) - ord("０")))
        else:
            out.append(c)
    return int("".join(out))


def normalise_date(v):
    """Coerce a value into 'YYYY.MM' where it looks like a date; otherwise pass through."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return f"{v.year}.{v.month:02d}"
    if isinstance(v, (int, float)):
        # Excel serial date (1900-system) — only convert if value is in plausible range
        if 30000 < v < 60000:
            base = datetime(1899, 12, 30)
            try:
                d = base + timedelta(days=float(v))
                return f"{d.year}.{d.month:02d}"
            except Exception:
                return str(v)
        return str(v)
    s = str(v).strip()
    # Already an ISO date string?
    m = re.match(r"^(\d{4})[-./](\d{1,2})", s)
    if m:
        return f"{m.group(1)}.{int(m.group(2)):02d}"
    return s


def dump_sheet(ws):
    """Return a dict keyed by category name."""
    repeated_blocks = defaultdict(list)  # name -> list of {label: value, _n: N}
    singletons = {}  # label -> value (for one-shot fields like 自我评价 / 居住地址)
    grouped = defaultdict(dict)  # 基本信息 / 链接 / 技能栈 etc. -> {label: value}

    cur_block = None        # dict in flight
    cur_cat = None
    cur_n = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = str(row[0]).strip() if row[0] is not None else ""
        b = str(row[1]).strip() if row[1] is not None else ""
        c = row[2] if len(row) > 2 else None

        if not a:
            # blank row: close any open repeated block
            if cur_block is not None:
                repeated_blocks[cur_cat].append(cur_block)
                cur_block, cur_cat, cur_n = None, None, None
            continue

        m = REPEATED_PATTERN.match(a)
        if m and m.group(1) in REPEATED_CATEGORIES:
            cat = m.group(1)
            n = fullwidth_to_int(m.group(2))
            new_block = (
                cur_block is None
                or cur_cat != cat
                or cur_n != n
                or b in cur_block
            )
            if new_block:
                if cur_block is not None:
                    repeated_blocks[cur_cat].append(cur_block)
                cur_block = {"_n": n}
                cur_cat = cat
                cur_n = n
            label = b
            value = c
            if any(k in label for k in ("时间", "Time", "Date", "时长")):
                value = normalise_date(value)
            else:
                value = "" if value is None else (
                    str(value).strip() if not isinstance(value, str) else value.strip()
                )
            cur_block[label] = value
        else:
            # singleton / grouped category — close any open repeated block
            if cur_block is not None:
                repeated_blocks[cur_cat].append(cur_block)
                cur_block, cur_cat, cur_n = None, None, None

            label = b
            value = c
            if value is None:
                value = ""
            elif isinstance(value, datetime):
                value = normalise_date(value)
            else:
                value = str(value).strip() if not isinstance(value, str) else value.strip()
            # Decide grouping: known multi-row categories get their own dict
            if a in {
                "基本信息", "居住地址", "工作意向", "链接",
                "技能栈", "技能爱好", "奖励情况",
                "自我介绍", "自我评价", "应聘理由", "作品附件",
            }:
                if label:
                    grouped[a][label] = value
                else:
                    grouped[a].setdefault("_value", value)
            else:
                # truly singleton row
                key = f"{a}::{label}" if label else a
                singletons[key] = value

    # close final block
    if cur_block is not None:
        repeated_blocks[cur_cat].append(cur_block)

    result = {}
    for cat, blocks in repeated_blocks.items():
        # sort by _n, but drop duplicates preferring the block with more fields
        seen_n = {}
        for blk in blocks:
            n = blk.get("_n", 0)
            if n not in seen_n or len(blk) > len(seen_n[n]):
                seen_n[n] = blk
        result[cat] = [seen_n[n] for n in sorted(seen_n.keys())]
    for cat, kv in grouped.items():
        result[cat] = kv
    if singletons:
        result["_singletons"] = singletons
    return result


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    out = {}
    for short, sheet_name in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  sheet not found: {sheet_name}", file=sys.stderr)
            continue
        out[short] = dump_sheet(wb[sheet_name])
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # Print a compact summary
    print(f"✅ dumped to {OUT_PATH}")
    for lang, data in out.items():
        cats = ", ".join(f"{k}({len(v) if isinstance(v, list) else len(v)})" for k, v in data.items() if k != "_singletons")
        print(f"  [{lang}] {cats}")


if __name__ == "__main__":
    main()
